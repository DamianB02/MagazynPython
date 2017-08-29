from django.contrib.auth import authenticate, login , logout
from django.contrib.auth.models import  User
from django.http import HttpResponseRedirect
from django.http.response import HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404
from django.template.context_processors import request
from django.utils import timezone
import re
import json
from towar.models import *
from django.core import serializers
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Table, TableStyle, Paragraph
from openpyxl import Workbook,load_workbook

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .serializers import *
from rest_framework.parsers import JSONParser
from rest_framework.decorators import api_view


@api_view(['GET', 'POST'])
def TowarList(request):
    
    if request.method == 'GET':
        towar = Towar.objects.all()
        serialiser = TowarSerializer(towar, many=True)
        return Response(serialiser.data)
        
    elif request.method == 'POST':
        towar = JSONParser().parse(request)
        serializer = TowarSerializer(data=towar)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET', 'PUT','DELETE'])
def towar_detail(request, pk):
  
    try:
        towar= Towar.objects.get(pk=pk)
    except Towar.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
  
    if request.method == 'GET':
        serializer = TowarSerializer(towar)
        return Response(serializer.data)
  
    elif request.method == 'PUT':
        data = JSONParser().parse(request)
        serializer = TowarSerializer(towar, data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
  
    elif request.method == 'DELETE':
        towar.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
      
#      

def index(request):
    lista_towarow = Towar.objects.all()
#     page = request.GET.get('page', 1)
#     paginator = Paginator(lista_towarow, 5)
#     try:
#         users = paginator.page(page)
#     except PageNotAnInteger:
#         users = paginator.page(1)
#     except EmptyPage:
#         users = paginator.page(paginator.num_pages)
    return render(request,'listatowarow.html',{'listatowarow': lista_towarow})


def dane_wykres(request):
    data = Log.objects.all()
    dane = serializers.serialize('json',data)
    return HttpResponse(dane, content_type="application/json");

def remove(request):
    if request.user.is_authenticated():
        id = request.POST.get("id");
        try:
            post = get_object_or_404(Towar,pk=id)
            commodity = post.name
            username = request.user
            action = "Usuniecie"
            id_commodity = id
            log = Log(username=username,action=action,commodity=commodity,id_commodity=id_commodity)
            log.save() 
            post.delete()     
        except Exception as e:
                print e
                return HttpResponse(e.message)
        return HttpResponse("OK")
    else:
        return HttpResponse("Nie jestes zalogowany")

def szczegoly_uzytkownika(request):
    if request.user.is_superuser:
        pk = request.GET.get('pk')
        if pk!=None:
            post = get_object_or_404(User,pk=pk)
            
            return render(request,'wszystkie_dane_uzytkownika.html',{'szczegoly': post})
        return HttpResponseRedirect("/")
    return HttpResponseRedirect("/")


def lista_uzytkownikow(request):
    if request.user.is_superuser:
        lista_uzytkownikow = User.objects.all()
    
        return render(request,'listauzytkownikow.html',{'listauzytkownikow': lista_uzytkownikow})
    else:
        return HttpResponseRedirect("/")

def zmiana_hasla(request):
    if request.user.is_authenticated():
        if request.method == "POST":
            try:
                user = request.user
                stare_haslo= request.POST.get("stareHaslo");
                if user.check_password(stare_haslo):  
                    user.set_password(request.POST.get("noweHaslo"))   
                    user.save()
                    u = authenticate(username=user.username, password=request.POST.get("noweHaslo"))
                    if u is not None:
                        login(request, u)
                        return HttpResponse("OK")
                else:
                    return HttpResponse("Niepoprawne haslo")
            except Exception as e:
                print e
                return HttpResponse(e.message)
                          
        return HttpResponse("OK")
    else:
        return HttpResponse("Nie jest zalogowany")
    
    
def form_logowanie(request):
    if not request.user.is_authenticated():
        return render(request,'logowanie.html')
    return HttpResponseRedirect("/")


def spr_logowanie(request):
    if not request.user.is_authenticated():
        if request.method == "POST":
            try:
                username = request.POST.get('login')
                password = request.POST.get('haslo')
                user = authenticate(username=username, password=password)
                if user is not None:
                    login(request, user)
                    return HttpResponse("OK")
                else:
                    return HttpResponse("Podane zle dane")
            except Exception as e:
                print e
                return HttpResponse(e.message)
                          
        
    else:
        return HttpResponse("Jestes zalogowany")

def zmiana_dane_uzytkownik(request):
    if request.user.is_authenticated():
        if request.method == "POST":
            try:
                user = request.user
                user.first_name = request.POST.get("imie")
                user.last_name = request.POST.get("nazwisko")
                user.email = request.POST.get("email")
                user.save()
            except Exception as e:
                print e
                return HttpResponse(e.message)
                          
        return HttpResponse("OK")
    else:
        return HttpResponse("Nie jest zalogowany")


def przesylanie_pdf(request):
    if request.user.is_authenticated():
        pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
        c = canvas.Canvas('plik.pdf',pagesize=A4)
        width, height = A4 
        user = User.objects.all()
        produkty = Towar.objects.all()
        dane={}
        usun=0
        dodaj=0
        akcja = Log.objects.all()
        licznik=0
        for k in user:
            usun=0
            dodaj=0
            for i in akcja:
                if k.username==i.username:
                    if i.action=="Usuniecie":
                        usun=usun+1
                        dane[k.username]={"usuniecie":usun,"dodanie":dodaj}
                    else:
                        dodaj=dodaj+1
                        dane[k.username]={"usuniecie":usun,"dodanie":dodaj}
            licznik=licznik+1
        minus=120 
        c.setFont("Arial", 14)
        c.drawString(40,(height-minus), "Tabela ile kto dodaje")
        minus += 40
        data=[]
        for key, value in dane.iteritems():
            data.append([str(key)+":","ilosc dodan:"+str(value["dodanie"]),"ilosc usuniec:"+str(value["usuniecie"])])
        
        tt=Table(data)
        w, h = tt.wrapOn(c,400,100)
        tt.drawOn(c, 50, (height-minus))   
        minus+=w
        c.setFont("Arial", 14)
        c.drawString(40,(height-minus), "Tabela towarow")
        minus += w+20
        tablicatowarow=[]
        for p in produkty:
            tablicatowarow.append([p.name,p.ilosc,p.osoba_wprowadzajaca])
        tt2 = Table(tablicatowarow)
        w, h = tt2.wrapOn(c,400,100)
        tt2.drawOn(c, 50, (height-minus))   
        c.showPage()
        c.save()
    return HttpResponseRedirect("/")



def import_export_xlsx(request):
    if request.user.is_superuser:          
            return render(request,'importexport.html')
    return HttpResponseRedirect("/")
  
  
  
  
def import_xlsx(request):
    if request.user.is_authenticated():
        tab = []
        i=0;
        ws = load_workbook(request.FILES.get("file"))
        for sheet in ws:
            for row in sheet:
                tab[i]
                print row[0].value
    return HttpResponse("OK")  

def export_xlsx(request):
    if request.user.is_authenticated():
        lista_towarow = Towar.objects.all();
        book = Workbook()
        
        towar = book.active
        
        towar.title = "Towar"
        
        print  book.sheetnames
        towar.append(["Nazwa towaru","Ilosc","Data dodania","Id kategorii","Osoba wprowadzajaca","Uwagi"])
        for row in lista_towarow:
            towar.append([row.name,row.ilosc,str(row.data_wprowadzenia),row.categoria_id,str(row.osoba_wprowadzajaca),row.uwagi])
        plik=request.POST.get("nazwaPliku")
        bla = book.create_sheet("dupa")
        print  book.sheetnames
        book.save(plik+'.xlsx')
        
    return HttpResponseRedirect("/")

def dodanie_nowego_towaru(request):
    if request.user.is_authenticated():
        if request.method == "POST":
            try:
                nazwa = request.POST.get("nazwaTowaru")
                ilosc = request.POST.get("ilosc")
                kategoria = request.POST.get("kategoria")
                uwagi = request.POST.get("uwagi")
                osoba_wprowadzajaca = request.user
                commodity = nazwa
                username = request.user
                action = "Dodanie"
                towar = Towar(name= nazwa,
                              ilosc= ilosc,
                              categoria_id= kategoria,
                              osoba_wprowadzajaca = osoba_wprowadzajaca,
                               uwagi= uwagi)
                towar.save()
                id_commodity = towar.id
                log = Log(username=username,action=action,commodity=commodity,id_commodity=id_commodity)
                log.save() 
            except Exception as e:
                print e
                return HttpResponse(e.message)
                          
        return HttpResponse("OK")
    else:
        return HttpResponse("Nie jest zalogowany")    
 
 
def lista_logi(request):
     logi = Log.objects.all()
     return render(request,'lista_logi.html',{'listalogi': logi})
 
    
def dane_uzytkownika(request):
       return render(request,'dane_uzytkownika.html')


def form_uzytkownik(request):
    if request.user.is_authenticated():
        return render(request,'edycja_uzytkownik.html')
    return HttpResponseRedirect("/")


def form_towar(request):
    if request.user.is_authenticated():
        lista_kategorie = Category.objects.all();
        return render(request,'dodanie_towaru.html',{'listakategorii': lista_kategorie})
    return HttpResponseRedirect("/")


def wyloguj(request):
    try:
        logout(request)
    except Exception as e:
        print e
        return HttpResponse(e.message)
    return HttpResponse("OK")
    
    